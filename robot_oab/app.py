from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
from time import sleep
import openpyxl

def main():
    number_oab = 133864

    # Configuração do WebDriver
    driver = configure_webdriver()

    try:
        # 1. Entrar no site
        navigate_to_website(driver)
        
        # 2. Preencher número da OAB e selecionar estado
        fill_oab_and_select_state(driver, number_oab)
        
        # 3. Clicar em pesquisar
        click_search_button(driver)
        
        # 4. Entrar em cada um dos processos
        process_links = find_process_links(driver)
        extract_and_store_process_data(driver, process_links)

    finally:
        driver.quit()

def configure_webdriver():
    driver = webdriver.Chrome()
    driver.maximize_window()
    return driver

def navigate_to_website(driver):
    driver.get('https://pje-consulta-publica.tjmg.jus.br/')
    sleep(10)

def fill_oab_and_select_state(driver, number_oab):
    oab_field = driver.find_element(By.XPATH, "//input[@id='fPP:Decoration:numeroOAB']")
    oab_field.send_keys(number_oab)

    state_dropdown = Select(driver.find_element(By.XPATH, "//select[@id='fPP:Decoration:estadoComboOAB']"))
    state_dropdown.select_by_visible_text('SP')

def click_search_button(driver):
    search_button = driver.find_element(By.XPATH, "//input[@id='fPP:searchProcessos']")
    search_button.click()
    sleep(10)

def find_process_links(driver):
    process_links = driver.find_elements(By.XPATH, '//b[@class="btn-block"]')
    return process_links

def extract_and_store_process_data(driver, process_links):
    workbook = openpyxl.load_workbook('data.xlsx')

    for process_link in process_links:
        process_link.click()
        sleep(10)
        windows = driver.window_handles
        driver.switch_to.window(windows[-1])
        driver.set_window_size(1920, 1080)

        # 6. Extrair o número do processo e a data de distribuição
        number_process = driver.find_elements(By.XPATH, "//div[@class='col-sm-12 ']")[0].text
        distribution_date = driver.find_elements(By.XPATH, "//div[@class='value col-sm-12 ']")[1].text

        # 7. Extrair e guardar todas as últimas movimentações
        process_movements = driver.find_elements(By.XPATH, "//div[@id='j_id132:processoEventoPanel']//tr[contains(@class, 'rich-table-row')]//td//div/div//span")
        movements_list = [movement.text for movement in process_movements]

        # 8. Guardar tudo em uma planilha no Excel e separados por processos
        try:
            process_page = workbook[number_process]
        except KeyError:
            process_page = workbook.create_sheet(number_process)

        process_page['A1'].value = "Process Number"
        process_page['B1'].value = "Distribution Date"
        process_page['C1'].value = "Movements"
        
        process_page['A2'].value = number_process
        process_page['B2'].value = distribution_date
        
        for index, row in enumerate(process_page.iter_rows(min_row=2, max_row=len(movements_list), min_col=3, max_col=3)):
            for cell in row:
                cell.value = movements_list[index]

        workbook.save('data.xlsx')
        driver.close()
        sleep(5)
        driver.switch_to.window(driver.window_handles[0])

if __name__ == "__main__":
    main()
    